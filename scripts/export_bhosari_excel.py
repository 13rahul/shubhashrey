#!/usr/bin/env python3
"""Build multi-sheet Excel workbook from Bhosari MIDC prospect CSVs."""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = DATA / "Bhosari_MIDC_Prospects_Bharatweld.xlsx"

SHEETS = [
    ("Prospects", "bhosari-midc-prospects.csv"),
    ("Hot_List", "bhosari-midc-hot-list.csv"),
    ("Deferred", "bhosari-midc-deferred.csv"),
    ("Parked", "bhosari-midc-parked.csv"),
    ("Lookup_Tiers", "lookup_tiers.csv"),
    ("Outreach_Log", "bhosari-midc-outreach-log.csv"),
]


def main():
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    hot_fill = PatternFill("solid", fgColor="FFF2CC")
    phone_ok_fill = PatternFill("solid", fgColor="C6EFCE")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    wb = Workbook()
    wb.remove(wb.active)

    # README first
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "Bhosari MIDC Prospects — Bharatweld / Shubhshrey Industries"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    instructions = [
        "",
        "How to use",
        "1. Start on Hot_List (or filter Prospects: priority_label=Hot AND outreach_status=Not contacted).",
        "2. Prefer verification_status=Phone-ok (green) — website-confirmed contacts.",
        "3. Key columns for dialing: company_name, plant_name, products_services, plot_address, phone_primary, email.",
        "4. For directory-sourced numbers, open google_maps_url, confirm unit still exists, then call.",
        "5. Pitch: boilers/fab → E7018+E6013; sheet/gates → E6013; maintenance/foundry → cutting + Mn hardfacing.",
        "6. Log every touch on Outreach_Log and update outreach_status on Prospects.",
        "",
        "Tabs",
        "Prospects — Phase 1 master (100 companies) with plant + products/services enrichment",
        "Hot_List — priority_label=Hot only",
        "Deferred — candidates beyond the frozen 100",
        "Parked — non-fit industries",
        "Lookup_Tiers — industry → electrode → pitch mapping",
        "Outreach_Log — call/WhatsApp/meeting log template",
        "",
        "Note: Older directory phones may be stale. Re-verify before bulk WhatsApp.",
        "Enriched 2026-09-03: SNEHA, Jaisons, MILKON, Sankalp Steeltech, MACHINESPACE, Super-Tech J-251/252.",
    ]
    for i, line in enumerate(instructions, start=2):
        cell = ws.cell(i, 1, line)
        if line in ("How to use", "Tabs"):
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 110

    for sheet_name, csv_name in SHEETS:
        path = DATA / csv_name
        with path.open(encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))
        ws = wb.create_sheet(sheet_name)
        if not rows:
            continue
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        if sheet_name in ("Prospects", "Hot_List") and len(rows) > 1:
            headers = rows[0]
            pri_i = headers.index("priority_label") + 1 if "priority_label" in headers else None
            ver_i = headers.index("verification_status") + 1 if "verification_status" in headers else None
            for r_idx in range(2, len(rows) + 1):
                if pri_i and ws.cell(r_idx, pri_i).value == "Hot":
                    ws.cell(r_idx, pri_i).fill = hot_fill
                if ver_i and ws.cell(r_idx, ver_i).value == "Phone-ok":
                    ws.cell(r_idx, ver_i).fill = phone_ok_fill

        for c_idx, header in enumerate(rows[0], start=1):
            max_len = len(str(header))
            for r_idx in range(2, min(len(rows) + 1, 40)):
                v = rows[r_idx - 1][c_idx - 1] if c_idx - 1 < len(rows[r_idx - 1]) else ""
                max_len = max(max_len, min(len(str(v)), 48))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max(12, max_len + 2), 42)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
