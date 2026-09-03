#!/usr/bin/env python3
"""Build Maharashtra MIDC master list for CRM + sales Excel.

Source of industrial area names: MIDC EODB 'Plot List With Trees' registry
(https://eodb.midcindia.org/TypeofTrees.aspx) — district-wise industrial areas.
Plus common sales aliases used in outreach (Bhosari, TTC pockets, etc.).

Outputs:
  - data/maharashtra-midcs.csv
  - data/Maharashtra_MIDCs_Electrode_Target_Map.xlsx
  - updates includes/geo_data.php MIDC function from the official names + aliases
"""
from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
GEO_PHP = ROOT / "includes" / "geo_data.php"

# Official EODB industrial areas: (district, area_name)
OFFICIAL: list[tuple[str, str]] = [
    # Ahilyanagar
    ("Ahilyanagar", "Ahilyanagar"),
    ("Ahilyanagar", "Jamkhed"),
    ("Ahilyanagar", "Newasa"),
    ("Ahilyanagar", "Rahuri"),
    ("Ahilyanagar", "Shrirampur"),
    ("Ahilyanagar", "Supa Parner"),
    # Akola
    ("Akola", "Akola"),
    ("Akola", "Akola (Central Government Growth Center)"),
    ("Akola", "Akot (Mini)"),
    ("Akola", "Balapur (Mini)"),
    ("Akola", "Barshi Takli (Mini)"),
    ("Akola", "Murtizapur"),
    ("Akola", "Patur (Mini)"),
    ("Akola", "Telhara (Mini)"),
    # Amravati
    ("Amravati", "Achalpur"),
    ("Amravati", "Addl. Amravati (Textile Park)"),
    ("Amravati", "Amravati"),
    ("Amravati", "Anjangaon (Mini)"),
    ("Amravati", "Bhatkuli (Mini)"),
    ("Amravati", "Chandur Railway (Mini)"),
    ("Amravati", "Daryapur (Mini)"),
    ("Amravati", "Dhamangaon"),
    ("Amravati", "Dharni (Mini)"),
    ("Amravati", "Morshi (Mini)"),
    ("Amravati", "Nandgaon-Khandeshwar (Mini)"),
    ("Amravati", "Tiwasa (Mini)"),
    ("Amravati", "Warud (Mini)"),
    # Beed
    ("Beed", "Ashti (Mini)"),
    ("Beed", "Beed"),
    ("Beed", "Dharur (Mini)"),
    ("Beed", "Majalgaon"),
    ("Beed", "Patoda (Mini)"),
    # Bhandara
    ("Bhandara", "Bhandara"),
    ("Bhandara", "Lakhandur"),
    ("Bhandara", "Mohadi (Mini)"),
    ("Bhandara", "Pauni (Mini)"),
    ("Bhandara", "Tumsar (Mini)"),
    # Buldhana
    ("Buldhana", "Buldhana (Mini)"),
    ("Buldhana", "Chikhali"),
    ("Buldhana", "Deoulgaon Raja (Mini)"),
    ("Buldhana", "Khamgaon"),
    ("Buldhana", "Lonar (Mini)"),
    ("Buldhana", "Malkapur"),
    ("Buldhana", "Mehekar (Mini)"),
    ("Buldhana", "Sangrampur (Mini)"),
    # Chandrapur
    ("Chandrapur", "Addl. Chandrapur"),
    ("Chandrapur", "Bhadravati (Major)"),
    ("Chandrapur", "Bhadravati (Mini)"),
    ("Chandrapur", "Chandrapur"),
    ("Chandrapur", "Chandrapur Tadali (CGGC)"),
    ("Chandrapur", "Chimur (Mini)"),
    ("Chandrapur", "Ghugus"),
    ("Chandrapur", "Gondpimpri (Mini)"),
    ("Chandrapur", "Mul"),
    ("Chandrapur", "Nagbhid"),
    ("Chandrapur", "Rajura (Mini)"),
    ("Chandrapur", "Sindewahi (Mini)"),
    ("Chandrapur", "Warora"),
    # Chhatrapati Sambhajinagar
    ("Chhatrapati Sambhajinagar", "Ambad"),
    ("Chhatrapati Sambhajinagar", "Chhatrapati Sambhajinagar (Estate)"),
    ("Chhatrapati Sambhajinagar", "Chikalthana"),
    ("Chhatrapati Sambhajinagar", "Khultabad (Mini)"),
    ("Chhatrapati Sambhajinagar", "Shendra Five Star"),
    ("Chhatrapati Sambhajinagar", "Vaijapur"),
    ("Chhatrapati Sambhajinagar", "Waluj"),
    # Dharashiv
    ("Dharashiv", "Addl. Dharashiv"),
    ("Dharashiv", "Bhoom (Mini)"),
    ("Dharashiv", "Dharashiv"),
    ("Dharashiv", "Kalamb (Mini)"),
    ("Dharashiv", "Umarga"),
    # Dhule
    ("Dhule", "Addl. Dhule"),
    ("Dhule", "Brahmanvel"),
    ("Dhule", "Dhule"),
    ("Dhule", "Nardana Ph-1"),
    ("Dhule", "Nardana Ph-2"),
    ("Dhule", "Shahade (Mini)"),
    # Gadchiroli
    ("Gadchiroli", "Aheri (Mini)"),
    ("Gadchiroli", "Dhanora"),
    ("Gadchiroli", "Gadchiroli"),
    ("Gadchiroli", "Kurkheda"),
    # Gondia
    ("Gondia", "Deori"),
    ("Gondia", "Gondia"),
    ("Gondia", "Goregaon (Mini)"),
    ("Gondia", "Morgaon Arjuni (Mini)"),
    ("Gondia", "Tiroda"),
    # Hingoli
    ("Hingoli", "Basmat (Mini)"),
    ("Hingoli", "Hingoli"),
    ("Hingoli", "Kalamnuri (Mini)"),
    # Jalgaon
    ("Jalgaon", "Addl. Jalgaon"),
    ("Jalgaon", "Bhusaval"),
    ("Jalgaon", "Chalisgaon"),
    ("Jalgaon", "Jalgaon (Estate)"),
    # Jalna
    ("Jalna", "Addl. Jalna Ph-1"),
    ("Jalna", "Addl. Jalna Ph-2"),
    ("Jalna", "Addl. Jalna Ph-3"),
    ("Jalna", "Ambad (Mini)"),
    ("Jalna", "Bhokardan"),
    ("Jalna", "Jafrabad (Mini)"),
    ("Jalna", "Jalna (Estate)"),
    ("Jalna", "Partur"),
    # Kolhapur
    ("Kolhapur", "Ajara"),
    ("Kolhapur", "Gadhinglaj"),
    ("Kolhapur", "Gokul-Shirgaon"),
    ("Kolhapur", "Halkarni"),
    ("Kolhapur", "Kagal-Hatkanangale Five Star"),
    ("Kolhapur", "Shiroli"),
    # Latur
    ("Latur", "Addl. Latur"),
    ("Latur", "Ahmedpur (Mini)"),
    ("Latur", "Ausa"),
    ("Latur", "Latur"),
    ("Latur", "Nilanga (Mini)"),
    # Mumbai City
    ("Mumbai City", "Marol"),
    ("Mumbai City", "Marol (Samruddhi Venture Park)"),
    # Nagpur
    ("Nagpur", "Bhivapur (Mini)"),
    ("Nagpur", "Butibori"),
    ("Nagpur", "Butibori Five Star"),
    ("Nagpur", "Butibori Ph-2"),
    ("Nagpur", "Hingna"),
    ("Nagpur", "Kalmeshwar"),
    ("Nagpur", "Kamti Kanhan"),
    ("Nagpur", "Katol"),
    ("Nagpur", "Kuhi (Mini)"),
    ("Nagpur", "Narkhed (Mini)"),
    ("Nagpur", "Parshivni (Mini)"),
    ("Nagpur", "Saoner"),
    ("Nagpur", "Umred"),
    # Nanded
    ("Nanded", "Bhokar"),
    ("Nanded", "Deglur"),
    ("Nanded", "Kandhar (Mini)"),
    ("Nanded", "Kinwat"),
    ("Nanded", "Krushnur (SEZ) (Pharma)"),
    ("Nanded", "Mudkhed"),
    ("Nanded", "Nanded"),
    # Nandurbar
    ("Nandurbar", "Navapur"),
    # Nashik
    ("Nashik", "Addl. Dindori"),
    ("Nashik", "Addl. Sinnar (SEZ)"),
    ("Nashik", "Dindori"),
    ("Nashik", "Malegaon (Textile Park)"),
    ("Nashik", "Nashik (IT Park)"),
    ("Nashik", "Nashik-Ambad"),
    ("Nashik", "Nashik-Satpur"),
    ("Nashik", "Peth"),
    ("Nashik", "Sinnar"),
    ("Nashik", "Surgana"),
    ("Nashik", "Vinchur"),
    ("Nashik", "Yeola"),
    # Palghar
    ("Palghar", "Tarapur"),
    # Parbhani
    ("Parbhani", "Gangakhed (Mini)"),
    ("Parbhani", "Jintur"),
    ("Parbhani", "Parbhani"),
    # Pune
    ("Pune", "Addl. Kurkumbh (Patas)"),
    ("Pune", "Baramati"),
    ("Pune", "Bhigwan"),
    ("Pune", "Chakan Ph-1"),
    ("Pune", "Chakan Ph-2"),
    ("Pune", "Chakan Ph-3"),
    ("Pune", "Chakan Ph-4"),
    ("Pune", "Indapur"),
    ("Pune", "Jejuri"),
    ("Pune", "Kharadi Knowledge Park"),
    ("Pune", "Kurkumbh"),
    ("Pune", "Pandare"),
    ("Pune", "Pimpri-Chinchwad"),
    ("Pune", "Rajiv Gandhi Infotech Park (Hinjawadi) Ph-1"),
    ("Pune", "Rajiv Gandhi Infotech Park (Hinjawadi) Ph-2"),
    ("Pune", "Rajiv Gandhi Infotech Park (Hinjawadi) Ph-3"),
    ("Pune", "Ranjangaon"),
    ("Pune", "Talawade Software Park"),
    ("Pune", "Talegaon"),
    ("Pune", "Talegaon Floriculture Park"),
    # Raigad
    ("Raigad", "Addl. Mahad"),
    ("Raigad", "Addl. Patalganga"),
    ("Raigad", "Mahad"),
    ("Raigad", "Nagothane"),
    ("Raigad", "Patalganga"),
    ("Raigad", "Roha"),
    ("Raigad", "Taloja"),
    ("Raigad", "Usar"),
    ("Raigad", "Vile-Bhagad"),
    # Ratnagiri
    ("Ratnagiri", "Dabhol"),
    ("Ratnagiri", "Dapoli"),
    ("Ratnagiri", "Gane-Khadpoli"),
    ("Ratnagiri", "Kherdi-Chiplun"),
    ("Ratnagiri", "Lote-Parshuram"),
    ("Ratnagiri", "Ratnagiri-Mirjole"),
    ("Ratnagiri", "Sadavali"),
    ("Ratnagiri", "Sangameshwar (Mini)"),
    # Sangli
    ("Sangli", "Addl. Palus (Wine Park)"),
    ("Sangli", "Islampur"),
    ("Sangli", "Jath (Mini)"),
    ("Sangli", "Kadegaon (Mini)"),
    ("Sangli", "Kavathe-Mahankal"),
    ("Sangli", "Palus (Mini)"),
    ("Sangli", "Sangli-Miraj"),
    ("Sangli", "Shalgaon-Bombalewadi"),
    ("Sangli", "Shirala"),
    ("Sangli", "Vita"),
    # Satara
    ("Satara", "Addl. Phaltan"),
    ("Satara", "Karad"),
    ("Satara", "Koregaon"),
    ("Satara", "Lonand"),
    ("Satara", "Mhaswad"),
    ("Satara", "Patan"),
    ("Satara", "Phaltan"),
    ("Satara", "Phaltan (SEZ)"),
    ("Satara", "Satara"),
    ("Satara", "Wai"),
    # Sindhudurg
    ("Sindhudurg", "Kudal"),
    # Solapur
    ("Solapur", "Barshi"),
    ("Solapur", "Chincholi"),
    ("Solapur", "Karmala"),
    ("Solapur", "Kurduwadi"),
    ("Solapur", "Mangalwedha"),
    ("Solapur", "Solapur"),
    ("Solapur", "Tembhurni"),
    # Thane
    ("Thane", "Addl. Ambernath"),
    ("Thane", "Addl. Ambernath Pale Ph-3"),
    ("Thane", "Addl. Murbad"),
    ("Thane", "Ambernath"),
    ("Thane", "Badlapur"),
    ("Thane", "Dombivli"),
    ("Thane", "Kalyan-Bhiwandi"),
    ("Thane", "Mira"),
    ("Thane", "Murbad"),
    ("Thane", "T.T.C."),
    ("Thane", "Thane"),
    # Wardha
    ("Wardha", "Deoli"),
    ("Wardha", "Hinganghat (Mini)"),
    ("Wardha", "Karanja (Mini)"),
    ("Wardha", "Samudrapur (Mini)"),
    ("Wardha", "Wardha"),
    # Washim
    ("Washim", "Malegaon (Mini)"),
    ("Washim", "Mangrulpir (Mini)"),
    ("Washim", "Manora (Mini)"),
    ("Washim", "Risod (Mini)"),
    ("Washim", "Washim"),
    # Yavatmal
    ("Yavatmal", "Addl. Yavatmal"),
    ("Yavatmal", "Darwha (Mini)"),
    ("Yavatmal", "Digras (Mini)"),
    ("Yavatmal", "Ghatanji (Mini)"),
    ("Yavatmal", "Kalamb (Mini)"),
    ("Yavatmal", "Kelapur Pandharkawada"),
    ("Yavatmal", "Mahagaon (Mini)"),
    ("Yavatmal", "Maregaon (Mini)"),
    ("Yavatmal", "Pusad"),
    ("Yavatmal", "Umarkhed (Mini)"),
    ("Yavatmal", "Wani"),
    ("Yavatmal", "Yavatmal"),
]

# Sales aliases used in CRM (not always separate official area names)
ALIASES: list[tuple[str, str]] = [
    ("Pune", "Bhosari MIDC"),
    ("Pune", "Pimpri MIDC"),
    ("Pune", "Chinchwad MIDC"),
    ("Pune", "Pimpri-Chinchwad MIDC"),
    ("Pune", "Chakan MIDC"),
    ("Pune", "Chakan Ph-5"),
    ("Pune", "Talegaon MIDC"),
    ("Pune", "Ranjangaon MIDC"),
    ("Pune", "Ranjangaon Ph-1"),
    ("Pune", "Ranjangaon Ph-3"),
    ("Pune", "Baramati MIDC"),
    ("Pune", "Baramati Ph-2"),
    ("Pune", "Kurkumbh MIDC"),
    ("Pune", "Hinjawadi Ph-1"),
    ("Pune", "Hinjawadi Ph-2"),
    ("Pune", "Hinjawadi Ph-3"),
    ("Pune", "Hinjawadi Ph-4"),
    ("Pune", "Khed SEZ (KEIPL)"),
    ("Satara", "Khandala Ph-1"),
    ("Satara", "Khandala Ph-2"),
    ("Satara", "Khandala Ph-3"),
    ("Thane", "TTC MIDC (Trans Thane Creek)"),
    ("Thane", "Mahape MIDC"),
    ("Thane", "Airoli MIDC"),
    ("Thane", "Rabale MIDC"),
    ("Thane", "Wagle Estate"),
    ("Thane", "Ambernath MIDC"),
    ("Thane", "Dombivli MIDC"),
    ("Raigad", "Taloja MIDC"),
    ("Raigad", "Taloja Ph-2"),
    ("Raigad", "Patalganga MIDC"),
    ("Raigad", "Patalganga Borivali"),
    ("Raigad", "Mahad MIDC"),
    ("Raigad", "Roha MIDC"),
    ("Raigad", "Vile Bhagad MIDC"),
    ("Palghar", "Boisar MIDC"),
    ("Palghar", "Tarapur MIDC"),
    ("Ratnagiri", "Lote Parshuram MIDC"),
    ("Chhatrapati Sambhajinagar", "Waluj MIDC"),
    ("Chhatrapati Sambhajinagar", "Chikalthana MIDC"),
    ("Chhatrapati Sambhajinagar", "Shendra MIDC"),
    ("Chhatrapati Sambhajinagar", "Shendra SEZ"),
    ("Chhatrapati Sambhajinagar", "Paithan MIDC"),
    ("Nashik", "Ambad MIDC (Nashik)"),
    ("Nashik", "Sinnar Ph-1"),
    ("Nashik", "Sinnar Ph-2"),
    ("Nashik", "Sinnar Ph-3"),
    ("Nashik", "Sinnar Ph-4"),
    ("Nagpur", "Butibori MIDC"),
    ("Nagpur", "Hingna MIDC (Nagpur)"),
    ("Kolhapur", "Kagal Hatkanangale Five Star"),
    ("Kolhapur", "Shiroli MIDC (Kolhapur)"),
    ("Sangli", "Sangli-Miraj-Kupwad MIDC"),
    ("Satara", "Satara MIDC"),
    ("Satara", "Karad MIDC"),
    ("Satara", "Phaltan MIDC"),
    ("Solapur", "Solapur MIDC"),
    ("Nandurbar", "Addl. Nandurbar (Bhaler)"),
    ("Mumbai City", "SEEPZ (SEZ)"),
    ("", "Other / Outside MIDC"),
    ("", "Non-MIDC Industrial Estate"),
]

REGION = {
    "Ahilyanagar": "North Maharashtra",
    "Akola": "Vidarbha",
    "Amravati": "Vidarbha",
    "Beed": "Marathwada",
    "Bhandara": "Vidarbha",
    "Buldhana": "Vidarbha",
    "Chandrapur": "Vidarbha",
    "Chhatrapati Sambhajinagar": "Marathwada",
    "Dharashiv": "Marathwada",
    "Dhule": "North Maharashtra",
    "Gadchiroli": "Vidarbha",
    "Gondia": "Vidarbha",
    "Hingoli": "Marathwada",
    "Jalgaon": "North Maharashtra",
    "Jalna": "Marathwada",
    "Kolhapur": "Western Maharashtra",
    "Latur": "Marathwada",
    "Mumbai City": "Mumbai Metro",
    "Mumbai Suburban": "Mumbai Metro",
    "Nagpur": "Vidarbha",
    "Nanded": "Marathwada",
    "Nandurbar": "North Maharashtra",
    "Nashik": "North Maharashtra",
    "Palghar": "Mumbai Metro",
    "Parbhani": "Marathwada",
    "Pune": "Pune",
    "Raigad": "Konkan / Raigad",
    "Ratnagiri": "Konkan",
    "Sangli": "Western Maharashtra",
    "Satara": "Western Maharashtra",
    "Sindhudurg": "Konkan",
    "Solapur": "Western Maharashtra",
    "Thane": "Mumbai Metro",
    "Wardha": "Vidarbha",
    "Washim": "Vidarbha",
    "Yavatmal": "Vidarbha",
}

# Names / patterns → High electrode demand (fab, auto, heavy eng, chemical plant)
HIGH_EXACT = {
    "Pimpri-Chinchwad",
    "Bhosari MIDC",
    "Pimpri MIDC",
    "Chinchwad MIDC",
    "Pimpri-Chinchwad MIDC",
    "Chakan Ph-1",
    "Chakan Ph-2",
    "Chakan Ph-3",
    "Chakan Ph-4",
    "Chakan Ph-5",
    "Chakan MIDC",
    "Ranjangaon",
    "Ranjangaon MIDC",
    "Ranjangaon Ph-1",
    "Ranjangaon Ph-3",
    "Talegaon",
    "Talegaon MIDC",
    "Baramati",
    "Baramati MIDC",
    "Baramati Ph-2",
    "Jejuri",
    "Kurkumbh",
    "Kurkumbh MIDC",
    "Addl. Kurkumbh (Patas)",
    "Indapur",
    "Pandare",
    "Taloja",
    "Taloja MIDC",
    "Taloja Ph-2",
    "Ambernath",
    "Ambernath MIDC",
    "Addl. Ambernath",
    "Addl. Ambernath Pale Ph-3",
    "Dombivli",
    "Dombivli MIDC",
    "T.T.C.",
    "TTC MIDC (Trans Thane Creek)",
    "Mahape MIDC",
    "Airoli MIDC",
    "Rabale MIDC",
    "Wagle Estate",
    "Patalganga",
    "Patalganga MIDC",
    "Addl. Patalganga",
    "Patalganga Borivali",
    "Mahad",
    "Mahad MIDC",
    "Addl. Mahad",
    "Nagothane",
    "Roha",
    "Roha MIDC",
    "Vile-Bhagad",
    "Vile Bhagad MIDC",
    "Tarapur",
    "Tarapur MIDC",
    "Boisar MIDC",
    "Nashik-Ambad",
    "Nashik-Satpur",
    "Ambad MIDC (Nashik)",
    "Sinnar",
    "Addl. Sinnar (SEZ)",
    "Sinnar Ph-1",
    "Sinnar Ph-2",
    "Sinnar Ph-3",
    "Sinnar Ph-4",
    "Dindori",
    "Addl. Dindori",
    "Vinchur",
    "Waluj",
    "Waluj MIDC",
    "Chikalthana",
    "Chikalthana MIDC",
    "Shendra Five Star",
    "Shendra MIDC",
    "Shendra SEZ",
    "Paithan MIDC",
    "Butibori",
    "Butibori MIDC",
    "Butibori Five Star",
    "Butibori Ph-2",
    "Hingna",
    "Hingna MIDC (Nagpur)",
    "Kalmeshwar",
    "Gokul-Shirgaon",
    "Shiroli",
    "Shiroli MIDC (Kolhapur)",
    "Kagal-Hatkanangale Five Star",
    "Kagal Hatkanangale Five Star",
    "Lote-Parshuram",
    "Lote Parshuram MIDC",
    "Jalna (Estate)",
    "Addl. Jalna Ph-1",
    "Addl. Jalna Ph-2",
    "Addl. Jalna Ph-3",
    "Amravati",
    "Addl. Amravati (Textile Park)",
    "Chandrapur",
    "Addl. Chandrapur",
    "Chandrapur Tadali (CGGC)",
    "Ghugus",
    "Solapur",
    "Solapur MIDC",
    "Chincholi",
    "Satara",
    "Satara MIDC",
    "Karad",
    "Karad MIDC",
    "Khandala Ph-1",
    "Khandala Ph-2",
    "Khandala Ph-3",
    "Lonand",
    "Sangli-Miraj",
    "Sangli-Miraj-Kupwad MIDC",
    "Islampur",
    "Ahilyanagar",
    "Supa Parner",
    "Dhule",
    "Nardana Ph-1",
    "Nardana Ph-2",
    "Jalgaon (Estate)",
    "Addl. Jalgaon",
    "Latur",
    "Addl. Latur",
    "Nanded",
}

LOW_PATTERNS = (
    "(Mini)",
    "Floriculture",
    "Wine Park",
    "Software Park",
    "Knowledge Park",
    "IT Park",
    "Samruddhi Venture",
)


def classify(name: str, district: str) -> tuple[str, str, str]:
    """Return (electrode_potential, sales_priority, typical_sectors)."""
    if name in ("Other / Outside MIDC", "Non-MIDC Industrial Estate"):
        return "Varies", "Case-by-case", "Any industrial buyer outside MIDC"
    if name in HIGH_EXACT:
        return (
            "High",
            "1 — Primary territory",
            "Heavy fab / auto / engineering / plant maintenance",
        )
    if any(p in name for p in LOW_PATTERNS):
        return "Low", "3 — Opportunistic", "Small workshops / niche / light industry"
    if district in {
        "Pune",
        "Thane",
        "Raigad",
        "Nashik",
        "Nagpur",
        "Chhatrapati Sambhajinagar",
        "Kolhapur",
        "Palghar",
        "Satara",
        "Sangli",
        "Solapur",
        "Ahilyanagar",
        "Jalna",
        "Amravati",
        "Chandrapur",
        "Ratnagiri",
    }:
        return "Medium", "2 — Expand after Tier-1", "General engineering / mixed manufacturing"
    return "Medium–Low", "3 — Opportunistic", "Mixed / smaller industrial base"


def php_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def update_geo_php(names: list[str]) -> None:
    text = GEO_PHP.read_text(encoding="utf-8")
    start = text.find("/**\n * Maharashtra MIDC / industrial areas")
    end = text.find("/**\n * @return list<string>\n */\nfunction shubh_districts_for_state")
    if start < 0 or end < 0:
        raise SystemExit("Could not locate MIDC function in geo_data.php")

    lines = [
        "/**",
        " * Maharashtra MIDC industrial areas (MIDC EODB registry) + sales aliases.",
        " * Rebuild via: python scripts/build_maharashtra_midcs.py",
        " *",
        " * @return list<string>",
        " */",
        "function shubh_maharashtra_midc_list(): array",
        "{",
        "    $list = [",
    ]
    for n in names:
        lines.append(f"        '{php_escape(n)}',")
    lines.extend(
        [
            "    ];",
            "",
            "    $list = array_values(array_unique($list));",
            "    natcasesort($list);",
            "    return array_values($list);",
            "}",
            "",
            "",
        ]
    )
    GEO_PHP.write_text(text[:start] + "\n".join(lines) + text[end:], encoding="utf-8")


def write_csv(rows: list[dict]) -> Path:
    import csv

    path = DATA / "maharashtra-midcs.csv"
    fields = [
        "midc_name",
        "district",
        "region",
        "electrode_potential",
        "sales_priority",
        "typical_sectors",
        "source",
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    return path


def write_excel(rows: list[dict]) -> Path | None:
    if Workbook is None:
        return None
    path = DATA / "Maharashtra_MIDCs_Electrode_Target_Map.xlsx"
    wb = Workbook()

    # README
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Maharashtra MIDC — Welding Electrode Target Map"
    ws["A1"].font = Font(bold=True, size=14)
    notes = [
        "",
        "Purpose: territory plan for Bharatweld / Shubhshrey Industries electrode sales.",
        "Official industrial-area names sourced from MIDC EODB Plot List With Trees.",
        "Maharashtra Industries Dept. cites 289+ MIDC areas; this sheet lists the EODB",
        "registry areas plus common sales aliases (Bhosari, Mahape, etc.).",
        "",
        "Electrode potential = likelihood of continuous electrode consumption",
        "(fab / heavy eng / auto / plant maintenance), not plot count alone.",
        "",
        "Suggested sales order (Pune first):",
        "  Bhosari / Pimpri-Chinchwad → Chakan → Talegaon → Ranjangaon → Baramati → Jejuri",
        "Then Mumbai/Raigad: Taloja → Ambernath → Patalganga → Mahad → Nagothane",
        "Then Nashik: Satpur → Ambad → Sinnar → Dindori",
        "Then Aurangabad belt: Waluj → Chikalthana → Shendra",
        "Then Nagpur: Hingna → Butibori",
        "",
        "Rebuild: python scripts/build_maharashtra_midcs.py",
    ]
    for i, line in enumerate(notes, start=2):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 100

    # All MIDCs
    all_ws = wb.create_sheet("All_MIDCs")
    headers = [
        "midc_name",
        "district",
        "region",
        "electrode_potential",
        "sales_priority",
        "typical_sectors",
        "source",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = all_ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            all_ws.cell(r, c, row[h])
    for col in range(1, len(headers) + 1):
        all_ws.column_dimensions[get_column_letter(col)].width = 28
    all_ws.auto_filter.ref = f"A1:G{len(rows)+1}"
    all_ws.freeze_panes = "A2"

    # High priority only
    hot = [r for r in rows if r["electrode_potential"] == "High"]
    hot_ws = wb.create_sheet("High_Potential")
    for col, h in enumerate(headers, 1):
        cell = hot_ws.cell(1, col, h)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = header_font
    for r, row in enumerate(hot, 2):
        for c, h in enumerate(headers, 1):
            hot_ws.cell(r, c, row[h])
    for col in range(1, len(headers) + 1):
        hot_ws.column_dimensions[get_column_letter(col)].width = 28
    hot_ws.auto_filter.ref = f"A1:G{len(hot)+1}"
    hot_ws.freeze_panes = "A2"

    # District summary
    from collections import Counter, defaultdict

    by_dist: dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        if not r["district"]:
            continue
        by_dist[r["district"]][r["electrode_potential"]] += 1

    sum_ws = wb.create_sheet("By_District")
    sum_headers = ["district", "region", "total_areas", "high", "medium", "low_or_other"]
    for col, h in enumerate(sum_headers, 1):
        cell = sum_ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for i, dist in enumerate(sorted(by_dist.keys()), 2):
        c = by_dist[dist]
        total = sum(c.values())
        high = c.get("High", 0)
        med = c.get("Medium", 0) + c.get("Medium–Low", 0)
        low = total - high - med
        sum_ws.cell(i, 1, dist)
        sum_ws.cell(i, 2, REGION.get(dist, ""))
        sum_ws.cell(i, 3, total)
        sum_ws.cell(i, 4, high)
        sum_ws.cell(i, 5, med)
        sum_ws.cell(i, 6, low)
    for col in range(1, 7):
        sum_ws.column_dimensions[get_column_letter(col)].width = 22
    sum_ws.freeze_panes = "A2"

    wb.save(path)
    return path


def main() -> None:
    rows: list[dict] = []
    seen: set[str] = set()

    for district, name in OFFICIAL + ALIASES:
        if name in seen:
            continue
        seen.add(name)
        pot, pri, sectors = classify(name, district)
        source = (
            "Sales alias"
            if any(a[1] == name for a in ALIASES)
            else "MIDC EODB registry"
        )
        rows.append(
            {
                "midc_name": name,
                "district": district,
                "region": REGION.get(district, ""),
                "electrode_potential": pot,
                "sales_priority": pri,
                "typical_sectors": sectors,
                "source": source,
            }
        )

    rows.sort(key=lambda r: (r["district"] or "zzz", r["midc_name"].lower()))

    DATA.mkdir(parents=True, exist_ok=True)
    csv_path = write_csv(rows)
    xlsx_path = write_excel(rows)

    names = sorted({r["midc_name"] for r in rows}, key=str.lower)
    update_geo_php(names)

    high = sum(1 for r in rows if r["electrode_potential"] == "High")
    print(f"MIDC rows: {len(rows)} (official+aliases)")
    print(f"High potential: {high}")
    print(f"Wrote {csv_path}")
    if xlsx_path:
        print(f"Wrote {xlsx_path}")
    else:
        print("openpyxl missing — CSV only; pip install openpyxl for Excel")
    print(f"Updated {GEO_PHP}")


if __name__ == "__main__":
    main()
